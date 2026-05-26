import os
import tempfile
import unittest
import warnings

import openpyxl
from lemana_parser.config import BASE_HEADERS, CONFIG
from lemana_parser.excel_writer import write_xlsx


class ExcelWriterTests(unittest.TestCase):
    def test_write_xlsx_creates_expected_headers_and_values(self):
        original_output_dir = CONFIG["output_dir"]
        original_output_filename = CONFIG["output_filename"]

        with tempfile.TemporaryDirectory() as tmpdir:
            try:
                CONFIG["output_dir"] = tmpdir
                CONFIG["output_filename"] = "test_result.xlsx"

                with warnings.catch_warnings():
                    warnings.filterwarnings(
                        "ignore",
                        message=r"datetime\.datetime\.utcnow\(\) is deprecated.*",
                        category=DeprecationWarning,
                    )
                    out_path = write_xlsx(
                        [
                            {
                                "status": "ok",
                                "error": "",
                                "article": "123456",
                                "url": "https://lemanapro.ru/catalogue/x/123456/",
                                "name": "Тестовый товар",
                                "price": "100,00",
                                "image": "https://img.example/p.webp",
                                "characteristics": {"Материал": "Металл"},
                            }
                        ],
                        ["Материал"],
                    )

                self.assertTrue(os.path.exists(out_path))

                with warnings.catch_warnings():
                    warnings.filterwarnings(
                        "ignore",
                        message=r"datetime\.datetime\.utcnow\(\) is deprecated.*",
                        category=DeprecationWarning,
                    )
                    workbook = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
                try:
                    sheet = workbook.active

                    headers = [
                        sheet.cell(row=1, column=i).value for i in range(1, len(BASE_HEADERS) + 2)
                    ]
                    self.assertEqual(headers, BASE_HEADERS + ["Материал"])
                    self.assertEqual(sheet["A2"].value, "ok")
                    self.assertEqual(sheet["C2"].value, "123456")
                    self.assertEqual(sheet["E2"].value, "Тестовый товар")
                    self.assertEqual(sheet["F2"].value, "100,00")
                    self.assertEqual(
                        sheet.cell(row=2, column=len(BASE_HEADERS) + 1).value,
                        "Металл",
                    )
                finally:
                    workbook.close()
            finally:
                CONFIG["output_dir"] = original_output_dir
                CONFIG["output_filename"] = original_output_filename


if __name__ == "__main__":
    unittest.main()
