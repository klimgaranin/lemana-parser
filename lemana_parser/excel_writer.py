"""
excel_writer.py — Запись результатов в .xlsx.
Один вызов write_xlsx() — красиво, с форматированием и гиперссылками.
"""

import logging
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from lemana_parser.config import BASE_HEADERS, CONFIG
from lemana_parser.models import Product

logger = logging.getLogger("excel_writer")

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_ZEBRA_FILL = PatternFill("solid", fgColor="EBF3FB")
_LINK_FONT = Font(color="0563C1", underline="single", size=9)
_CELL_FONT = Font(size=9)
_BORDER_SIDE = Side(style="thin", color="D0D0D0")
_CELL_BORDER = Border(
    left=_BORDER_SIDE,
    right=_BORDER_SIDE,
    top=_BORDER_SIDE,
    bottom=_BORDER_SIDE,
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _write_cell(
    ws,
    col_idx: dict[str, int],
    row_i: int,
    fill: PatternFill | None,
    col_name: str,
    value: object,
    is_link: bool = False,
) -> None:
    ci = col_idx.get(col_name)
    if ci is None:
        return
    cell = ws.cell(row=row_i, column=ci, value=value)
    cell.border = _CELL_BORDER
    cell.alignment = _LEFT
    if fill:
        cell.fill = fill
    if is_link and value:
        cell.hyperlink = value
        cell.font = _LINK_FONT
    else:
        cell.font = _CELL_FONT


def write_xlsx(products: list[Product], char_keys: list[str]) -> str:
    os.makedirs(CONFIG["output_dir"], exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = CONFIG["output_filename"].replace(".xlsx", f"_{ts}.xlsx")
    out_path = os.path.join(CONFIG["output_dir"], fname)

    headers = BASE_HEADERS + list(char_keys)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результат"

    # ── Заголовки ─────────────────────────────────────────────────────────────
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _CELL_BORDER

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    # ── Данные ────────────────────────────────────────────────────────────────
    for row_i, p in enumerate(products, start=2):
        is_zebra = row_i % 2 == 0
        fill = _ZEBRA_FILL if is_zebra else None

        row_values = [
            ("Статус", p.get("status", "ok"), False),
            ("Ошибка", p.get("error", ""), False),
            ("Артикул ЛМ", p.get("article", ""), False),
            ("ССЫЛКА", p.get("url", ""), True),
            ("Наименование товара", p.get("name", ""), False),
            ("Цена на сайте", p.get("price", ""), False),
            ("Ссылка на картинку", p.get("image", ""), True),
        ]
        for col_name, value, is_link in row_values:
            _write_cell(ws, col_idx, row_i, fill, col_name, value, is_link)

        for key, val in (p.get("characteristics") or {}).items():
            _write_cell(ws, col_idx, row_i, fill, key, val)

    # ── Ширина колонок ────────────────────────────────────────────────────────
    COL_WIDTHS = {
        "Статус": 16,
        "Ошибка": 45,
        "Артикул ЛМ": 12,
        "ССЫЛКА": 40,
        "Наименование товара": 45,
        "Цена на сайте": 14,
        "Ссылка на картинку": 45,
    }
    for col_i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = COL_WIDTHS.get(h, 20)

    wb.save(out_path)
    wb.close()
    logger.info("✅ Excel сохранён: %s (%d строк)", out_path, len(products))
    return out_path
